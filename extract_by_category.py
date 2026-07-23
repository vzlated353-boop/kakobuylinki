import openpyxl
import csv
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = 'd:/fansbuy/finsbuy-main/S1-Fansbuy(1).xlsx'
OUTPUT_DIR = 'd:/fansbuy/finsbuy-main/category_csv'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Target columns in order
TARGET_HEADERS = ['产品名', '产品图片链接', '产品id', '产品购买链接', '价格', '二级分类', '一级分类', '品牌']

# Sheet -> category name mapping
# Each sheet IS a category
SHEET_MAP = {
    '衣服': 'clothing',
    '裤子': 'pants',
    '鞋子': 'shoe',
    '手表': 'watches',
    '配饰': 'ACC',
}

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

for sheet_name, category_name in SHEET_MAP.items():
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        print(f"[SKIP] {sheet_name}: no data")
        continue
    
    header = rows[0]
    print(f"\n{'='*50}")
    print(f"Sheet: {sheet_name} -> {category_name}")
    print(f"Headers: {header}")
    
    # Determine column indices for target fields
    # Most sheets have: 产品名(0), 产品图片链接(1), 产品id(2), 产品购买链接(3), 价格(4), 二级分类(5), 一级分类(6), 品牌(7)
    # Some sheets may have None header for brand column
    
    output_rows = []
    brand_col_idx = None
    
    # Find brand column (could be index 7 or 8)
    for i, h in enumerate(header):
        if h and '品牌' in str(h):
            brand_col_idx = i
            break
    
    for row in rows[1:]:
        if not row or not row[0]:  # skip empty rows
            continue
        
        product_name = str(row[0]).strip() if row[0] else ''
        if not product_name:
            continue
        
        image_url = str(row[1]).strip() if row[1] and not str(row[1]).startswith('#') else ''
        product_id = row[2] if row[2] else ''
        purchase_url = str(row[3]).strip() if row[3] and not str(row[3]).startswith('FANSBUY') else ''
        price = str(row[4]).strip() if row[4] else ''
        secondary_cat = str(row[5]).strip() if row[5] else ''
        primary_cat = str(row[6]).strip() if row[6] else category_name
        
        # Get brand
        brand = ''
        if brand_col_idx is not None and brand_col_idx < len(row):
            brand = str(row[brand_col_idx]).strip() if row[brand_col_idx] else ''
        
        output_rows.append([
            product_name, image_url, product_id, purchase_url,
            price, secondary_cat, primary_cat, brand
        ])
    
    # Write CSV
    csv_path = os.path.join(OUTPUT_DIR, f'{category_name}.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(TARGET_HEADERS)
        writer.writerows(output_rows)
    
    print(f"  -> Written {len(output_rows)} rows to {csv_path}")

# Handle sheet '22' - it has different structure, no brand/category info
ws22 = wb['22']
rows22 = list(ws22.iter_rows(values_only=True))
print(f"\n{'='*50}")
print(f"Sheet: 22 (no category info, {len(rows22)-1} data rows)")
print(f"Headers: {rows22[0] if rows22 else 'N/A'}")

wb.close()

print(f"\nDone! Files saved to {OUTPUT_DIR}")
print("Files:", os.listdir(OUTPUT_DIR))
