import openpyxl

wb = openpyxl.load_workbook('d:/fansbuy/finsbuy-main/S1-Fansbuy(1).xlsx', read_only=True, data_only=True)
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    rows = list(ws.iter_rows(max_row=5, values_only=True))
    if rows:
        # Print header
        print("Headers:", rows[0])
        # Print first few data rows
        for i, row in enumerate(rows[1:], 1):
            print(f"Row {i}: {row}")
    
    # Count total rows
    total = ws.max_row
    print(f"Total rows (including header): {total}")

wb.close()
